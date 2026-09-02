"""Generate and verify the concise human-facing Excel workbook."""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from .markdown import DatasheetError
from .model import DOWNLOAD_SIZE_RE, Candidate, Criteria, first_token


WORKBOOK_SHEETS = ["Datasheet", "Fields", "Technical", "Totals"]
DATASHEET_COLUMNS = [
    "Candidate",
    "Tier",
    "Modality",
    "Samples",
    "Download size",
    "Publication",
    "Metadata check",
    "Required-fields check",
    "Raw-data check",
    "Metadata access",
    "Raw-data access",
    "Metadata evidence",
    "Raw-data evidence",
    "Recommendation",
    "Human decision",
    "Access and contacts",
    "Metadata route",
    "Raw-data route",
]
STATUS_COLUMNS = [
    "Metadata check",
    "Required-fields check",
    "Raw-data check",
    "Metadata access",
    "Raw-data access",
    "Metadata evidence",
    "Raw-data evidence",
    "Recommendation",
    "Human decision",
]
URL_RE = re.compile(r"https?://[^\s<>()\[\]\"'`]+")
CITATION_RE = re.compile(r"\s*\[S\d+\]")
CELL_LIMIT = 32_000
BRIEF_SUFFIX = " … see candidate file"

COLORS = {
    "pass": ("C6EFCE", "006100"),
    "accept": ("C6EFCE", "006100"),
    "contents inspected": ("C6EFCE", "006100"),
    "direct": ("C6EFCE", "006100"),
    "present": ("C6EFCE", "006100"),
    "fail": ("FFC7CE", "9C0006"),
    "reject": ("FFC7CE", "9C0006"),
    "unavailable": ("FFC7CE", "9C0006"),
    "absent": ("FFC7CE", "9C0006"),
    "unknown": ("FFEB9C", "9C5700"),
    "claimed": ("FFEB9C", "9C5700"),
    "route verified": ("FFEB9C", "9C5700"),
    "on request": ("FFEB9C", "9C5700"),
    "off-repository": ("FFEB9C", "9C5700"),
    "pending": ("FFEB9C", "9C5700"),
    "not checked": ("FFEB9C", "9C5700"),
}


def _clean(value: str) -> str:
    return " ".join(CITATION_RE.sub("", value).split())


def _brief(value: str, limit: int) -> str:
    cleaned = _clean(value)
    if len(cleaned) <= limit:
        return cleaned
    return cleaned[: limit - len(BRIEF_SUFFIX)].rstrip(" ,;:") + BRIEF_SUFFIX


def _tier(value: str) -> str:
    cleaned = _clean(value)
    match = re.match(r"(?:tier\s*)?([^\s—–,;]+)", cleaned, re.I)
    return match.group(1) if match else "unknown"


def _evidence(value: str) -> str:
    cleaned = _clean(value).casefold()
    for token in ("contents inspected", "route verified", "claimed", "unknown"):
        if cleaned.startswith(token):
            return token
    return "unknown"


def _route(value: str) -> tuple[str, str | None]:
    urls = list(dict.fromkeys(url.rstrip(".,);") for url in URL_RE.findall(value)))
    if len(urls) == 1:
        return urls[0], urls[0]
    if len(urls) > 1:
        return "\n".join(urls), None
    return _clean(value), None


def _sample_count(candidate: Candidate) -> int:
    match = re.search(r"\b(\d[\d,]*)\b", candidate.summary.get("Sample count", ""))
    return int(match.group(1).replace(",", "")) if match else 0


def _download_size(value: str) -> str:
    cleaned = _clean(value)
    if cleaned.casefold().startswith("unknown"):
        return "unknown"
    match = DOWNLOAD_SIZE_RE.match(cleaned)
    return re.sub(r"\s+", " ", match.group("size")).strip() if match else "unknown"


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _paint(cell, token: str, Font, PatternFill) -> None:
    color = COLORS.get(token.casefold())
    if color:
        background, foreground = color
        cell.fill = PatternFill("solid", fgColor=background)
        cell.font = Font(color=foreground, bold=True)


def _style_sheet(sheet, widths: list[int], Font, PatternFill, Alignment) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="374151")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.auto_filter.ref = sheet.dimensions


def build_workbook(datasheet_dir: str, candidates: list[Candidate], criteria: Criteria) -> Path:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise DatasheetError(
            "build requires openpyxl in the skill runtime; it is not available in this Python environment"
        ) from exc

    candidates = sorted(candidates, key=lambda c: (_clean(c.summary.get("Tier", "")), c.ident))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Datasheet"
    sheet.append(DATASHEET_COLUMNS)

    for candidate in candidates:
        metadata_route, metadata_link = _route(candidate.summary["Metadata route"])
        raw_route, raw_link = _route(candidate.summary["Raw-data route"])
        row = [
            candidate.ident,
            _tier(candidate.summary["Tier"]),
            _brief(candidate.summary["Modality"], 120),
            _sample_count(candidate),
            _download_size(candidate.summary["Download size"]),
            _brief(candidate.summary["Publication"], 160),
            candidate.check_result("Metadata"),
            candidate.check_result("Required fields"),
            candidate.check_result("Raw data"),
            first_token(candidate.summary["Metadata access"]),
            first_token(candidate.summary["Raw-data access"]),
            _evidence(candidate.summary["Metadata evidence"]),
            _evidence(candidate.summary["Raw-data evidence"]),
            first_token(candidate.summary["Recommendation"]),
            first_token(candidate.human_decision),
            _brief(candidate.summary["Access and contacts"], 200),
            metadata_route,
            raw_route,
        ]
        sheet.append(row)
        row_index = sheet.max_row
        for name in STATUS_COLUMNS:
            column = DATASHEET_COLUMNS.index(name) + 1
            _paint(sheet.cell(row_index, column), str(sheet.cell(row_index, column).value or ""), Font, PatternFill)
        for name, link in (("Metadata route", metadata_link), ("Raw-data route", raw_link)):
            if link:
                cell = sheet.cell(row_index, DATASHEET_COLUMNS.index(name) + 1)
                cell.hyperlink = link
                cell.font = Font(color="0563C1", underline="single")

    _style_sheet(
        sheet,
        [22, 16, 30, 24, 20, 38, 18, 20, 18, 18, 18, 20, 20, 18, 18, 42, 38, 38],
        Font,
        PatternFill,
        Alignment,
    )
    sheet.freeze_panes = "B2"

    fields_sheet = workbook.create_sheet("Fields")
    field_names = [field.name for field in criteria.fields if "ignore" not in field.priority]
    field_headers = ["Field", "Priority", *[candidate.ident for candidate in candidates]]
    fields_sheet.append(field_headers)
    for name in field_names:
        priority = next(
            (field.priority for field in criteria.fields if field.name.casefold() == name.casefold()),
            "observed",
        )
        values = [name, priority]
        statuses = []
        for candidate in candidates:
            row = next((item for item in candidate.fields if item.name.casefold() == name.casefold()), None)
            if row is None:
                values.append("—")
                statuses.append("not checked")
            else:
                detail = ", ".join(part for part in (row.coverage, row.storage, row.level) if part)
                values.append(("absent — " if row.status.casefold() == "absent" else "") + detail)
                statuses.append(row.status.casefold())
        fields_sheet.append(values)
        for offset, status in enumerate(statuses, start=3):
            _paint(fields_sheet.cell(fields_sheet.max_row, offset), status, Font, PatternFill)
    _style_sheet(fields_sheet, [28, 16, *([28] * len(candidates))], Font, PatternFill, Alignment)
    fields_sheet.freeze_panes = "C2"

    technical_sheet = workbook.create_sheet("Technical")
    technical_headers = ["Item", *[candidate.ident for candidate in candidates]]
    technical_sheet.append(technical_headers)
    for item in criteria.technical_items:
        values = [item]
        for candidate in candidates:
            row = next((row for row in candidate.technical if row.item.casefold() == item.casefold()), None)
            values.append(_brief(row.value, 360) if row else "—")
        technical_sheet.append(values)
    _style_sheet(technical_sheet, [32, *([30] * len(candidates))], Font, PatternFill, Alignment)
    technical_sheet.freeze_panes = "B2"

    totals = workbook.create_sheet("Totals")
    totals.append(["Group", "Value", "Candidates", "Samples"])
    groupers = [
        ("Tier", lambda candidate: _tier(candidate.summary["Tier"])),
        ("Recommendation", lambda candidate: first_token(candidate.summary["Recommendation"])),
        ("Metadata check", lambda candidate: candidate.check_result("Metadata")),
        ("Required-fields check", lambda candidate: candidate.check_result("Required fields")),
        ("Raw-data check", lambda candidate: candidate.check_result("Raw data")),
        ("Raw-data access", lambda candidate: first_token(candidate.summary["Raw-data access"])),
    ]
    for group_name, key in groupers:
        buckets: dict[str, list[Candidate]] = defaultdict(list)
        for candidate in candidates:
            buckets[key(candidate) or "—"].append(candidate)
        for value in sorted(buckets, key=str.casefold):
            group = buckets[value]
            totals.append([group_name, value, len(group), sum(_sample_count(candidate) for candidate in group)])
    _style_sheet(totals, [24, 42, 14, 14], Font, PatternFill, Alignment)
    totals.freeze_panes = "A2"

    out = Path(datasheet_dir) / "datasheet.xlsx"
    try:
        workbook.save(out)
        reopened = load_workbook(out)
    except OSError as exc:
        raise DatasheetError(f"cannot write or reopen {out}: {exc}") from exc

    problems = []
    if reopened.sheetnames != WORKBOOK_SHEETS:
        problems.append(f"sheets are {reopened.sheetnames}; expected {WORKBOOK_SHEETS}")
    main = reopened["Datasheet"]
    if [cell.value for cell in main[1]] != DATASHEET_COLUMNS:
        problems.append("Datasheet headers do not match the review contract")
    if main.max_row - 1 != len(candidates):
        problems.append(f"Datasheet has {main.max_row - 1} candidates; expected {len(candidates)}")
    present = {main.cell(row, 1).value for row in range(2, main.max_row + 1)}
    for candidate in candidates:
        if candidate.ident not in present:
            problems.append(f"candidate missing from Datasheet: {candidate.ident}")
    if reopened["Fields"].max_row - 1 != len(field_names):
        problems.append("Fields row count does not match candidate records")
    if reopened["Technical"].max_row - 1 != len(criteria.technical_items):
        problems.append("Technical row count does not match criteria.md")
    if reopened["Totals"].max_row < 2:
        problems.append("Totals sheet is empty")
    for worksheet in reopened.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and len(cell.value) > CELL_LIMIT:
                    problems.append(f"{worksheet.title}!{cell.coordinate} is near Excel's text limit")
    if problems:
        raise DatasheetError("workbook verification failed: " + "; ".join(problems))
    return out
